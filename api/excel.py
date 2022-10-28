import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

from .models import Element


def insert_cells(ws: Worksheet, ws_row: int, cells: dict):
    """Insert cell value on row of `ws_row`, getted by column letter from `cells`"""
    for column in (ws.iter_cols(min_row=ws_row, min_col=1, max_col=8, max_row=ws_row)):
        for cell in column:
            letter = cell.column_letter
            if not cells.get(cell.column_letter):
                continue
            cell.value = cells[letter]["value"]
            cell.alignment = cells[letter].get("alignment", None)

    ws_row += 1

    return ws, ws_row


def foreman(project):
    stages = project["stages"]
    ws1_row = 1
    ws2_row = 1
    ws3_row = 1

    ws1_index = 1
    ws2_index = 1
    ws3_index = 1

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Список работ (Бригадир)"
    ws2 = wb.create_sheet("Список материалов (Бригадир)")
    ws3 = wb.create_sheet("Список общий (Бригадир)")

    for stage in stages:
        ws1.merge_cells(f"A{ws3_row}:I{ws3_row}")
        ws2.merge_cells(f"A{ws3_row}:I{ws3_row}")
        ws3.merge_cells(f"A{ws3_row}:I{ws3_row}")
        ws1[f"A{ws3_row}"] = ws2[f"A{ws3_row}"] = ws3[f"A{ws3_row}"] = f"Этап {stage['order']}. {stage['title']}"
        ws1[f"A{ws3_row}"].alignment = Alignment(horizontal="center")
        ws2[f"A{ws3_row}"].alignment = Alignment(horizontal="center")
        ws3[f"A{ws3_row}"].alignment = Alignment(horizontal="center")
        ws1_row += 1
        ws2_row += 1
        ws3_row += 1

        col_headers_work = {
            "B": {"value": "Наименование"},
            "C": {"value": "Описание"},
            "D": {"value": "Кол-во"},
            "E": {"value": "ед-изм"}
        }
        col_headers = {**cells_work,
                "F": {"value": "Кол-во"},
                "G": {"value": "доп ед-изм"},
                "H": {"value": "Вес"},
                "I": {"value": "Объем"}
        }

        ws1, ws1_row = insert_cells(ws1, ws1_row, col_headers_work)
        ws2, ws2_row = insert_cells(ws2, ws2_row, col_headers)
        ws3, ws3_row = insert_cells(ws3, ws3_row, col_headers)

        constructions = stage["constructions"]
        for count_construction, construction in enumerate(constructions, start=1):
            cells = {
                "A": {"value": f"{count_construction}. Конструкция"},
                "B": {"value": construction["title"]},
                "D": {"value": construction["measure"]}
            }
            ws1, ws1_row = insert_cells(ws1, ws1_row, cells)
            ws2, ws2_row = insert_cells(ws2, ws2_row, cells)
            ws3, ws3_row = insert_cells(ws3, ws3_row, cells)

            elements = construction["elements"]
            for element in elements:
                cells_work = {
                        "A": {
                            "value": None,
                            "alignment": Alignment(horizontal="right")
                        },
                        "B": {
                            "value": element["title"]
                        },
                        "C": {
                            "value": element["original_title"]
                        },
                        "D": {
                            "value": element["count"],
                            "alignment": Alignment(horizontal="right")
                        },
                        "E": {
                            "value": element["measure"]
                        }
                    }
                cells = {**cells_work,
                        "F": {
                            "value": element["count"] * element["conversion_rate"],
                            "alignment": Alignment(horizontal="right")
                        },
                        "G": {
                            "value": element["second_measure"]
                        },
                        "H": {
                            "value": element["weight"]
                        },
                        "I": {
                            "value": element["volume"]
                        }
                    }

                if element["type"] == Element.Type.JOB:
                    cells["A"]["value"] = f"{count_construction}.{ws1_index}"
                    ws1, ws1_row = insert_cells(ws1, ws1_row, cells_work)
                    ws1_index += 1
                elif element["type"] == Element.Type.MATERIAL:
                    cells["A"]["value"] = f"{count_construction}.{ws2_index}"
                    ws2, ws2_row = insert_cells(ws2, ws2_row, cells)
                    ws2_index += 1

                cells["A"]["value"] = f"{count_construction}.{ws3_index}"
                ws3, ws3_row = insert_cells(ws3, ws3_row, cells)
                ws3_index +=1 

            count_construction += 1

    return wb


def purchaser(project):
    stages = project["stages"]
    ws1_row = 1
    ws2_row = 1

    ws1_index = 1
    ws2_index = 1

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Список материалов по этапам (Закупщик)"
    ws2 = wb.create_sheet("Список материалов общий (Закупщик)")

    for stage_index, stage in enumerate(stages):
        ws1.merge_cells(f"A{ws1_row}:F{ws1_row}")
        ws1[f"A{ws1_row}"] = f"Этап {stage['order']}. {stage['title']}"
        ws1[f"A{ws1_row}"].alignment = Alignment(horizontal="center")
        ws1_row += 1

        cells = {
            "B": {"value": "Наименование"},
            "C": {"value": "Описание"},
            "D": {"value": "Кол-во"},
            "E": {"value": "ед-изм"},
            "F": {"value": "Кол-во"},
            "G": {"value": "доп ед-изм"},
            "H": {"value": "Цена/ед"},
            "I": {"value": "Сумма"},
            "J": {"value": "Вес"},
            "K": {"value": "Объем"},
            "L": {"value": "Размер"}
        }
        ws1, ws1_row = insert_cells(ws1, ws1_row, cells)
        if stage_index == 0:
            ws2, ws2_row = insert_cells(ws2, ws2_row, cells)

        constructions = stage["constructions"]
        for count_construction, construction in enumerate(constructions, start=1):
            cells = {
                "A": {"value": f"{count_construction}. Конструкция"},
                "B": {"value": construction["title"]},
                "D": {"value": construction["measure"]}
            }
            ws1, ws1_row = insert_cells(ws1, ws1_row, cells)

            elements = construction["elements"]
            for element in elements:
                cells = {
                    "A": {
                        "value": None,
                        "alignment": Alignment(horizontal="right")
                    },
                    "B": {
                        "value": element["title"]
                    },
                    "C": {
                        "value": element["original_title"]
                    },
                    "D": {
                        "value": element["count"],
                        "alignment": Alignment(horizontal="right")
                    },
                    "E": {
                        "value": element["measure"]
                    },
                    "F": {
                        "value": element["count"] * element["conversion_rate"],
                    },
                    "G": {
                        "value": element["second_measure"]
                    },
                    "H": {
                        "value": element["cost"]
                    },
                    "I": {
                        "value": None
                    },
                    "J": {
                        "value": element["weight"]
                    },
                    "K": {
                        "value": element["volume"]
                    },
                    "L": {
                        "value": element["dimension"]
                    }
                }
                if element["type"] == Element.Type.MATERIAL:
                    cells["A"]["value"] = f"{count_construction}.{ws1_index}"
                    cells["I"]["value"] = f"=D{ws1_row}*H{ws1_row}"
                    ws1, ws1_row = insert_cells(ws1, ws1_row, cells)
                    ws1_index += 1

                cells["A"]["value"] = f"{count_construction}.{ws2_index}"
                cells["I"]["value"] = f"=D{ws2_row}*H{ws2_row}"
                ws2, ws2_row = insert_cells(ws2, ws2_row, cells)
                ws2_index += 1

            count_construction += 1

    return wb


def estimate(project):
    def sum_total_price(ws, ws_row, ws_price_cells, word: str):
        ws.merge_cells(f"A{ws_row}:F{ws_row}")
        ws[f"A{ws_row}"] = word
        ws[f"H{ws_row}"] = f"=SUM({';'.join(ws_price_cells)})"

        return ws, ws_row

    stages = project["stages"]
    ws1_row = 1
    ws2_row = 1

    ws1_index = 1

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Смета"
    ws2 = wb.create_sheet("Смета(сокр)")

    ws1_stage_price_cells = []
    ws2_stage_price_cells = []

    for stage in stages:
        ws1.merge_cells(f"A{ws1_row}:H{ws1_row}")
        ws2.merge_cells(f"A{ws2_row}:H{ws2_row}")
        ws1[f"A{ws1_row}"] = ws2[f"A{ws2_row}"] = f"Этап {stage['order']}. {stage['title']}"
        ws1[f"A{ws1_row}"].alignment = Alignment(horizontal="center")
        ws2[f"A{ws2_row}"].alignment = Alignment(horizontal="center")
        ws1_row += 1
        ws2_row += 1

        cells = {
            "B": {"value": "Наименование"},
            "C": {"value": "Описание"},
            "D": {"value": "Кол-во"},
            "E": {"value": "ед-изм"},
            "F": {"value": "Цена/ед"},
            "G": {"value": "Сумма"},
            "H": {"value": "Итого"}
        }
        cells_full = {**cells, "H": {"value": "Итого"}}

        ws1, ws1_row = insert_cells(ws1, ws1_row, cells_full)
        ws2, ws2_row = insert_cells(ws2, ws2_row, cells)

        ws1_construction_price_cells = []
        ws2_construction_price_cells = []

        constructions = stage["constructions"]
        for count_construction, construction in enumerate(constructions, start=1):
            cells = {
                "A": {"value": f"{count_construction}. Конструкция"},
                "B": {"value": construction["title"]},
                "C": {"value": construction["count"]},
                "D": {"value": construction["measure"]},
                "F": {"value": None}
            }
            cells_full = {**cells, 
                "C": {"value": None},
                "D": {"value": construction["count"]},
                "E": {"value": construction["measure"]},
                "F": {"value": None},
                "G": {"value": f"=SUM(G{ws1_row + 1}:G{ws1_row + len(construction['elements'])})"}
            }

            ws1, ws1_row = insert_cells(ws1, ws1_row, cells_full)
            cells["F"]["value"] = f"=SUM(F{ws2_row + 1}:F{ws2_row + len(construction['elements'])})"
            ws2, ws2_row = insert_cells(ws2, ws2_row, cells)

            ws1_construction_price_cells.append(f"G{ws1_row-1}")
            ws2_construction_price_cells.append(f"F{ws2_row-1}")

            elements_price = 0

            elements = construction["elements"]
            for element in elements:
                cells = {
                    "A": {
                        "value": f"{count_construction}.{ws1_index}",
                        "alignment": Alignment(horizontal="right")
                    },
                    "B": {
                        "value": element["title"]
                    },
                    "C": {
                        "value": element["original_title"],
                    },
                    "D": {
                        "value": element["count"],
                    },
                    "E": {
                        "value": element["measure"]
                    },
                    "F": {
                        "value": element["cost"]
                    },
                    "G": {
                        "value": f"=D{ws1_row}*F{ws1_row}"
                    }
                }
                ws1, ws1_row = insert_cells(ws1, ws1_row, cells)

                ws1_index += 1

                elements_price += element['cost'] * element['count']

            ws2[f"F{ws2_row - 1}"] = elements_price
            count_construction += 1

        ws1, ws1_row = sum_total_price(ws1, ws1_row, ws1_construction_price_cells, "Итого")
        ws1_stage_price_cells.append(f"H{ws1_row}")
        ws1_row += 1

        ws2, ws2_row = sum_total_price(ws2, ws2_row, ws2_construction_price_cells, "Итого")
        ws2_stage_price_cells.append(f"H{ws2_row}")
        ws2_row += 1

    ws1, ws1_row = sum_total_price(ws1, ws1_row, ws1_stage_price_cells, "Всего")
    ws2, ws2_row = sum_total_price(ws2, ws2_row, ws2_stage_price_cells, "Всего")

    return wb

def export(elements: List[Element]):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1_row = 4

    current_category = None

    ws1.merge_cells("D1:E1")
    ws1["D1"] = "Коммерческие единицы"
    ws1.merge_cells("F1:G1")
    ws1["F1"] = "Строительные единицы"
    ws1.merge_cells("H1:I1")
    ws1["H1"] = "Себестоимость коммерческие единицы"
    ws1.merge_cells("J1:K1")
    ws1["J1"] = "Себестоимость строительные единицы"

    ws1["C2"] = "Наименование"
    ws1["D2"] = "Ст-ть"
    ws1["E2"] = "ед"
    ws1["F2"] = "Ст-ть"
    ws1["G2"] = "ед"
    ws1["H2"] = "Ст-ть"
    ws1["I2"] = "ед"
    ws1["J2"] = "Ст-ть"
    ws1["K2"] = "ед"

    for element in elements:
        if not hasattr(element, "subcategory"):
            continue

        if not current_category or current_category != element.subcategory.title:
            if current_category:
                # Add blank rows
                ws1_row += 2

            current_category = element.subcategory.title

            ws1.merge_cells(f"C{ws1_row}:K{ws1_row}")
            ws1[f"C{ws1_row}"].fill = PatternFill(fgColor="FCE89C", fill_type = "solid")
            ws1[f"C{ws1_row}"] = element.subcategory.title

        ws1_row += 1

        cells = {
            "C": {"value": element.title},
            "D": {"value": element.price * element.conversion_rate},
            "E": {"value": element.measure},
            "F": {"value": element.price},
            "G": {"value": element.second_measure},
            "H": {"value": element.cost * element.conversion_rate},
            "I": {"value": element.measure},
            "J": {"value": element.cost},
            "K": {"value": element.second_measure},
        }
        ws1, _ = insert_cells(ws1, ws1_row, cells)

    return wb

def q_import(wb: Workbook):
    ws1 = wb.active
    elements = []
    column_map = {
        "A": "title",
        "B": "measure",
        "C": "second_measure",
        "D": "cost",
        "E": "price",
        "F": "type",
        "G": "dimension",
        "H": "conversion_rate",
        "I": "weight",
        "J": "volume",
        "K": "parent_category",
        "L": "category",
        "M": "subcategory"
    }

    errors = []
    error_messages = {
        "not_exists": "Объекта с id %(value)s не существует!"
    }

    required_fields = [
        f.name for f in Element._meta.get_fields()
        if not getattr(f, 'blank', False) is True and
        not isinstance(f, models.ForeignObjectRel)
    ]

    row: List[Cell]
    for row in ws1.iter_rows(min_row=2):
        data = {}

        if all(v.value is None for v in row):
            # Skip empty rows
            continue

        for cell in row:
            field_name = column_map[cell.column_letter]
            field = Element._meta.get_field(field_name)

            if field.name not in required_fields and not cell.value:
                continue

            try:
                if field.is_relation and not field.related_model.objects.filter(id=cell.value).exists():
                    # Check if related model exists with cell.value as id
                    raise ValidationError(
                        error_messages["not_exists"],
                        code="invalid_related",
                        params={"value": int(cell.value)}
                    )

                field.clean(cell.value, field.model)
            except ValidationError as e:
                error = f"{cell.coordinate} - {e.messages[0]}"
                errors.append(error)
                continue

            data[field.attname] = cell.value

        elements.append(Element(**data))

    if errors:
        return None, errors

    return elements, None